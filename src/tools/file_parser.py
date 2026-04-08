"""Local file parser tool — fast schema inspection without sandbox round-trip."""

from __future__ import annotations

import io
import os
from typing import Any

import streamlit as st
from langchain_core.tools import tool


MAX_PREVIEW_ROWS = 100


def _parse_csv(file_bytes: bytes, filename: str, sep: str = ",") -> dict[str, Any]:
    import pandas as pd

    df = pd.read_csv(io.BytesIO(file_bytes), nrows=MAX_PREVIEW_ROWS, sep=sep)
    total_rows = file_bytes.count(b'\n') - 1  # minus header

    return {
        "type": "csv",
        "filename": filename,
        "columns": list(df.columns),
        "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
        "preview_rows": len(df),
        "total_rows": total_rows,
        "file_size_mb": round(len(file_bytes) / (1024 * 1024), 2),
        "preview": df.head(10).to_string(),
    }


def _parse_tsv(file_bytes: bytes, filename: str) -> dict[str, Any]:
    return _parse_csv(file_bytes, filename, sep='\t')


def _excel_numfmt_to_strftime(number_format: str) -> str:
    """Convert Excel number format string to Python strftime format."""
    if not number_format or number_format == 'General':
        return '%Y-%m-%d'

    # Strip time part (e.g., 'YYYY-MM-DD HH:MM:SS' → 'YYYY-MM-DD')
    nf = number_format.lower().strip()
    nf = nf.split(' hh')[0].split(' h:')[0].strip()

    # Common Excel date format mappings (date part only)
    mappings = [
        ('mm/dd/yyyy', '%m/%d/%Y'),
        ('m/d/yyyy', '%m/%d/%Y'),
        ('dd/mm/yyyy', '%d/%m/%Y'),
        ('d/m/yyyy', '%d/%m/%Y'),
        ('dd.mm.yyyy', '%d.%m.%Y'),
        ('d.m.yyyy', '%d.%m.%Y'),
        ('yyyy-mm-dd', '%Y-%m-%d'),
        ('yyyy/mm/dd', '%Y/%m/%d'),
        ('dd-mm-yyyy', '%d-%m-%Y'),
        ('mm-dd-yyyy', '%m-%d-%Y'),
        ('d/m/yy', '%d/%m/%y'),
        ('m/d/yy', '%m/%d/%y'),
        ('dd/mm/yy', '%d/%m/%y'),
        ('mm/dd/yy', '%m/%d/%y'),
        ('mm-dd-yy', '%m-%d-%y'),
        ('dd-mm-yy', '%d-%m-%y'),
        ('yy-mm-dd', '%y-%m-%d'),
    ]

    for excel_fmt, py_fmt in mappings:
        if excel_fmt in nf:
            return py_fmt

    return '%Y-%m-%d'


def _detect_date_format_from_strings(series) -> dict | None:
    """Detect date format from a pandas Series containing string date values."""
    import re
    samples = series.dropna().head(5).tolist()
    if not samples:
        return None

    str_samples = [str(s) for s in samples]

    formats_to_try = [
        ('%m/%d/%Y', r'^\d{1,2}/\d{1,2}/\d{4}$'),
        ('%d.%m.%Y', r'^\d{1,2}\.\d{1,2}\.\d{4}$'),
        ('%Y-%m-%d', r'^\d{4}-\d{2}-\d{2}'),
        ('%d/%m/%Y', r'^\d{1,2}/\d{1,2}/\d{4}$'),
        ('%d-%m-%Y', r'^\d{1,2}-\d{1,2}-\d{4}$'),
    ]

    for fmt, pattern in formats_to_try:
        if all(re.match(pattern, s[:10]) for s in str_samples if s != 'NaT'):
            return {"format": fmt, "samples": str_samples[:3]}

    return {"format": "unknown", "samples": str_samples[:3]}


def _parse_excel(file_bytes: bytes, filename: str) -> dict[str, Any]:
    import pandas as pd
    from openpyxl import load_workbook

    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    sheet_names = xls.sheet_names

    # Load workbook with openpyxl to read cell number formats
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    sheets_info = {}
    for sheet in sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet, nrows=MAX_PREVIEW_ROWS)

        date_columns = {}
        for col in df.columns:
            dtype = str(df[col].dtype)
            if dtype in ('datetime64[ns]', 'datetime64[us]'):
                # Read actual Excel cell format via openpyxl
                col_idx = list(df.columns).index(col)
                py_fmt = '%Y-%m-%d'  # default
                try:
                    ws = wb[sheet]
                    # Check first few data rows for number format
                    for row in ws.iter_rows(min_row=2, max_row=6,
                                            min_col=col_idx + 1,
                                            max_col=col_idx + 1):
                        for cell in row:
                            if cell.number_format and cell.number_format != 'General':
                                py_fmt = _excel_numfmt_to_strftime(cell.number_format)
                                break
                        if py_fmt != '%Y-%m-%d':
                            break
                except Exception:
                    pass

                samples = df[col].dropna().head(3).apply(
                    lambda x: x.strftime(py_fmt)
                ).tolist()
                date_columns[str(col)] = {"format": py_fmt, "samples": samples}

            elif dtype == 'object':
                sample = df[col].dropna().head(3)
                if len(sample) > 0:
                    import re
                    date_like = any(
                        re.match(r'\d{1,2}[/.\-]\d{1,2}[/.\-]\d{4}', str(v))
                        for v in sample
                    )
                    if date_like:
                        info = _detect_date_format_from_strings(df[col])
                        if info:
                            date_columns[str(col)] = info

        sheet_entry = {
            "columns": list(df.columns),
            "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
            "preview_rows": len(df),
            "preview": df.head(5).to_string(),
        }
        if date_columns:
            sheet_entry["date_columns"] = date_columns

        sheets_info[sheet] = sheet_entry

    wb.close()

    return {
        "type": "excel",
        "filename": filename,
        "sheets": sheet_names,
        "sheet_count": len(sheet_names),
        "file_size_mb": round(len(file_bytes) / (1024 * 1024), 2),
        "sheets_info": sheets_info,
    }


def _parse_json(file_bytes: bytes, filename: str) -> dict[str, Any]:
    import pandas as pd

    try:
        df = pd.read_json(io.BytesIO(file_bytes), lines=True, nrows=MAX_PREVIEW_ROWS)
    except ValueError:
        df = pd.read_json(io.BytesIO(file_bytes))
    return {
        "type": "json",
        "filename": filename,
        "columns": list(df.columns),
        "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
        "total_rows": len(df),
        "file_size_mb": round(len(file_bytes) / (1024 * 1024), 2),
        "preview": df.head(10).to_string(),
    }


def _parse_pdf(file_bytes: bytes, filename: str) -> dict[str, Any]:
    import pdfplumber

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        pages_info = []
        text_preview = ""
        tables_found = 0

        for i, page in enumerate(pdf.pages[:5]):  # First 5 pages
            page_text = page.extract_text() or ""
            page_tables = page.extract_tables() or []
            tables_found += len(page_tables)
            pages_info.append({
                "page": i + 1,
                "text_length": len(page_text),
                "tables": len(page_tables),
            })
            if i == 0:
                text_preview = page_text[:500]

        return {
            "type": "pdf",
            "filename": filename,
            "total_pages": len(pdf.pages),
            "pages_previewed": len(pages_info),
            "tables_found": tables_found,
            "file_size_mb": round(len(file_bytes) / (1024 * 1024), 2),
            "pages_info": pages_info,
            "text_preview": text_preview,
        }


PARSERS = {
    ".csv": _parse_csv,
    ".tsv": _parse_tsv,
    ".xlsx": _parse_excel,
    ".xls": _parse_excel,
    ".xlsm": _parse_excel,
    ".json": _parse_json,
    ".pdf": _parse_pdf,
}


def make_parse_file_tool(uploaded_files: list | None = None):
    """Factory: create the parse_file tool with access to uploaded files.

    Args:
        uploaded_files: List of uploaded file objects (from st.file_uploader).
                       If None, falls back to st.session_state (for backwards compat).
    """
    # Capture uploaded_files in closure to avoid st.session_state access in agent thread
    _files = uploaded_files if uploaded_files is not None else st.session_state.get("uploaded_files", [])

    @tool
    def parse_file(filename: str) -> str:
        """Parse an uploaded file and return its schema summary.

        Returns column names, data types, row count, file size, and a preview.
        Use this FIRST before analyzing any file — it's fast (runs locally).

        Args:
            filename: Name of the uploaded file (e.g., "sales.xlsx")
        """
        # Use captured _files from closure (thread-safe, no st.session_state access)
        uploaded_files = _files
        target = None
        for f in uploaded_files:
            if f.name == filename:
                target = f
                break

        if target is None:
            available = [f.name for f in uploaded_files] if uploaded_files else []
            return f"File '{filename}' not found. Available files: {available}"

        ext = os.path.splitext(filename)[1].lower()
        parser = PARSERS.get(ext)
        if parser is None:
            return f"Unsupported file type: {ext}. Supported: {list(PARSERS.keys())}"

        try:
            target.seek(0)
            data = target.getvalue()
            result = parser(data, filename)
            size_mb = len(data) / (1024 * 1024)
            output = str(result)

            # Add explicit next-step instruction to prevent ls/parse_file loops
            output += f"\n\n{'='*70}\n"
            output += "✅ PARSE BAŞARILI. SONRAKI ADIM:\n\n"
            output += "❌ YAPMA: ls, cat, os.listdir, parse_file tekrar çağırma\n"
            output += "✅ YAP:\n"
            output += f"1. DÜŞÜNCE yaz: 'Schema alındı. Dosya /home/sandbox/{filename}. Şimdi pd.read_excel ile okuyacağım.'\n"
            output += "2. execute() çağır:\n"
            output += f"   df = pd.read_excel('/home/sandbox/{filename}')\n"
            output += "   print(f'✅ {{len(df)}} satır yüklendi')\n"

            # Add date format instructions if date columns detected
            if isinstance(result, dict):
                for sheet_info in result.get('sheets_info', {}).values():
                    date_cols = sheet_info.get('date_columns', {})
                    if date_cols:
                        output += "\n\n📅 TARİH FORMATI TALİMATI (ZORUNLU):\n"
                        for col_name, col_info in date_cols.items():
                            fmt = col_info.get('format', '%Y-%m-%d')
                            samples = col_info.get('samples', [])
                            output += f"   Kolon: '{col_name}' → Excel formatı: '{fmt}' → Örnekler: {samples}\n"
                        output += "\n   ⚠️ ZORUNLU KURALLAR:\n"
                        output += "   1. EKRANA YAZARKEN (print): strftime(fmt) kullan → '00:00:00' OLMAMALI\n"
                        output += "      display_df = df.copy()\n"
                        for col_name, col_info in date_cols.items():
                            fmt = col_info.get('format', '%Y-%m-%d')
                            output += f"      display_df['{col_name}'] = display_df['{col_name}'].dt.strftime('{fmt}')\n"
                        output += "      print(display_df)  # TÜM kolonlar, tarihler orijinal formatta\n"
                        output += "   2. EXCEL'E YAZARKEN (to_excel): datetime olarak BIRAK, strftime YAPMA\n"
                        output += "   3. FİLTRELEME: pd.to_datetime(user_input, format=fmt) kullan\n"

            output += f"\n⚠️ SONRAKİ TOOL: execute (BAŞKA HİÇBİR TOOL ÇAĞIRMA!)\n"
            output += "{'='*70}"

            if size_mb >= 40:
                output += (
                    f"\n\n⚠️ BÜYÜK DOSYA ({size_mb:.1f} MB ≥ 40MB) — DUCKDB STRATEJİSİ ZORUNLU. "
                    "pandas ile pd.read_excel() KULLANMA (çok yavaş/bellek sorunu). "
                    "Doğru strateji: "
                    "① df = pd.read_excel(path) ile CSV'ye çevir: df.to_csv('/home/sandbox/temp.csv', index=False); del df "
                    "② duckdb.sql(\"SELECT ... FROM read_csv_auto('/home/sandbox/temp.csv')\").df() "
                    "Threshold: file_size_mb >= 40 → DuckDB, < 40 → pandas."
                )
            return output
        except Exception as e:
            return f"Error parsing '{filename}': {e}"

    return parse_file
