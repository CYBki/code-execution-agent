"""Download file tool — fetch a file from OpenSandbox and offer as Streamlit download."""

from __future__ import annotations

import io
import logging
import os

from langchain_core.tools import tool
from src.sandbox.manager import OpenSandboxBackend, SANDBOX_HOME

from src.tools.artifact_store import get_store

logger = logging.getLogger(__name__)


def _clean_excel_dates(content: bytes) -> bytes:
    """Strip time part from date-only columns in Excel files.

    If a datetime column has ALL values at midnight (00:00:00),
    convert it to date-only so Excel doesn't show the time part.
    """
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = load_workbook(io.BytesIO(content))
        modified = False

        for ws in wb.worksheets:
            for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
                # Check if all non-None values in this column are midnight datetimes
                from datetime import datetime
                cells_with_dates = [
                    c for c in col
                    if isinstance(c.value, datetime)
                ]
                if not cells_with_dates:
                    continue

                all_midnight = all(
                    c.value.hour == 0 and c.value.minute == 0 and c.value.second == 0
                    for c in cells_with_dates
                )
                if all_midnight:
                    for c in cells_with_dates:
                        c.value = c.value.date()  # datetime → date (no time)
                        c.number_format = 'YYYY-MM-DD'
                    modified = True

        if modified:
            buf = io.BytesIO()
            wb.save(buf)
            wb.close()
            return buf.getvalue()

        wb.close()
        return content
    except Exception as e:
        logger.debug("Excel date cleanup skipped: %s", e)
        return content


def make_download_file_tool(backend: OpenSandboxBackend, session_id: str = ""):
    """Factory: create the download_file tool bound to an OpenSandbox backend."""

    @tool
    def download_file(file_path: str) -> str:
        """Make a file from the sandbox available for the user to download.

        Use this AFTER generating a file (PDF report, Excel output, CSV export, etc.)
        in the sandbox. The file will appear as a download button in the chat.

        Args:
            file_path: Absolute path to the file in the sandbox, e.g. '/home/sandbox/report.pdf'
        """
        ALLOWED_PREFIX = SANDBOX_HOME + "/"
        if not file_path.startswith(ALLOWED_PREFIX):
            return f"❌ Only files under {ALLOWED_PREFIX} can be downloaded."

        filename = os.path.basename(file_path)

        try:
            responses = backend.download_files([file_path])
            resp = responses[0] if responses and len(responses) > 0 else None

            if resp and resp.content and not getattr(resp, "error", None):
                MAX_DOWNLOAD_MB = 50
                if len(resp.content) > MAX_DOWNLOAD_MB * 1024 * 1024:
                    return f"❌ File too large ({len(resp.content) // 1024 // 1024}MB). Max {MAX_DOWNLOAD_MB}MB."

                file_content = resp.content
                # Clean date-only columns in Excel files (remove 00:00:00)
                if filename.lower().endswith(('.xlsx', '.xlsm')):
                    file_content = _clean_excel_dates(file_content)

                # Thread-safe per-session store — no st.session_state access from agent thread
                get_store(session_id).add_download(file_content, filename, file_path)
                size_kb = len(resp.content) / 1024
                logger.info("File ready for download: %s (%.1f KB)", filename, size_kb)
                return f"✅ '{filename}' ready for download ({size_kb:.1f} KB)."
            else:
                error_info = getattr(resp, "error", "no response") if resp else "no response"
                return f"❌ Could not download '{file_path}': {error_info}"
        except Exception as e:
            logger.error("Failed to download %s: %s", file_path, e, exc_info=True)
            return f"❌ Error downloading '{file_path}': {e}"

    return download_file
